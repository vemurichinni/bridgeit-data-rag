#!/usr/bin/env python3
"""
census_files.py — inventory a document archive for RAG planning.

Walks one or more root folders and records, per file: path, extension, size,
modified year, SHA-256 (for duplicate detection) and, where cheap to detect,
parsing hazards: scanned PDFs (no text layer), encrypted PDFs, legacy Office
formats (.doc/.xls/.ppt), multi-sheet workbooks, zero-byte files.

Outputs (into --out, default ./census_out):
  files.csv            one row per file
  files_summary.json   counts and bytes by extension / year / hazard
  samples.csv          the "20 nasty files per type" list for parser testing

Usage:
  python census_files.py /archive/projects /archive/shared --out census_out
  python census_files.py /archive --no-hash          # faster, skips dedupe
  python census_files.py /archive --samples 30       # more samples per type

Only the standard library is required. If `pypdf` and `openpyxl` are installed
the PDF/Excel hazard checks run; otherwise those columns are left blank.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    import warnings
    warnings.filterwarnings("ignore")
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover
    PdfReader = None

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None

# Extension -> family used for summaries and sampling
FAMILIES = {
    "pdf": "pdf",
    "doc": "word-legacy", "docx": "word", "rtf": "word", "odt": "word",
    "xls": "excel-legacy", "xlsx": "excel", "xlsm": "excel", "csv": "excel", "ods": "excel",
    "ppt": "ppt-legacy", "pptx": "ppt",
    "txt": "text", "md": "text", "log": "text",
    "msg": "email", "eml": "email", "mbox": "email", "pst": "email-pst",
    "png": "image", "jpg": "image", "jpeg": "image", "tif": "image", "tiff": "image", "gif": "image", "bmp": "image",
    "zip": "archive", "rar": "archive", "7z": "archive", "tar": "archive", "gz": "archive",
    "java": "code", "sql": "code", "xml": "code", "js": "code", "ts": "code", "html": "code",
    "properties": "code", "yml": "code", "yaml": "code", "json": "code", "cs": "code", "py": "code",
    "jar": "binary", "war": "binary", "ear": "binary", "class": "binary", "dll": "binary", "exe": "binary",
    "bak": "db-dump", "mdf": "db-dump", "dmp": "db-dump", "sav": "db-dump",
}
SKIP_DIRS = {".git", "node_modules", "target", "build", "dist", ".idea", ".svn", "__pycache__", "$RECYCLE.BIN"}


def sha256(path: Path, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(chunk), b""):
            h.update(block)
    return h.hexdigest()


def pdf_hazards(path: Path) -> dict:
    """Return {'pdf_pages', 'pdf_encrypted', 'pdf_scanned'} using pypdf if available."""
    out = {"pdf_pages": "", "pdf_encrypted": "", "pdf_scanned": ""}
    if PdfReader is None:
        return out
    try:
        r = PdfReader(str(path))
        out["pdf_encrypted"] = "yes" if r.is_encrypted else "no"
        if r.is_encrypted:
            try:
                r.decrypt("")
            except Exception:
                return out
        n = len(r.pages)
        out["pdf_pages"] = n
        # Sample up to 3 pages; if none yield text, it is almost certainly a scan.
        idx = sorted({0, n // 2, n - 1}) if n else []
        text_len = 0
        for i in idx:
            try:
                text_len += len((r.pages[i].extract_text() or "").strip())
            except Exception:
                pass
        out["pdf_scanned"] = "yes" if n and text_len < 20 * len(idx) else "no"
    except Exception as e:  # corrupt / unsupported
        out["pdf_encrypted"] = f"error:{type(e).__name__}"
    return out


def xlsx_hazards(path: Path) -> dict:
    out = {"xlsx_sheets": "", "xlsx_max_rows": "", "xlsx_has_formulas": ""}
    if openpyxl is None:
        return out
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
        sheets = wb.sheetnames
        out["xlsx_sheets"] = len(sheets)
        max_rows, formulas = 0, False
        for ws in wb.worksheets:
            max_rows = max(max_rows, ws.max_row or 0)
            if not formulas:
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 200)):
                    if any(isinstance(c.value, str) and c.value.startswith("=") for c in row):
                        formulas = True
                        break
        out["xlsx_max_rows"] = max_rows
        out["xlsx_has_formulas"] = "yes" if formulas else "no"
        wb.close()
    except Exception as e:
        out["xlsx_sheets"] = f"error:{type(e).__name__}"
    return out


def walk(roots: list[str], do_hash: bool, follow_links: bool):
    for root in roots:
        root_p = Path(root).expanduser().resolve()
        if not root_p.exists():
            print(f"!! root not found: {root}", file=sys.stderr)
            continue
        for dirpath, dirnames, filenames in os.walk(root_p, followlinks=follow_links):
            dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS]
            for name in filenames:
                p = Path(dirpath) / name
                try:
                    st = p.stat()
                except OSError:
                    continue
                ext = p.suffix.lower().lstrip(".")
                rec = {
                    "root": str(root_p),
                    "path": str(p),
                    "name": name,
                    "ext": ext,
                    "family": FAMILIES.get(ext, "other" if ext else "no-extension"),
                    "size_bytes": st.st_size,
                    "modified": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d"),
                    "year": datetime.fromtimestamp(st.st_mtime).year,
                    "depth": len(p.relative_to(root_p).parts),
                    "top_folder": p.relative_to(root_p).parts[0] if len(p.relative_to(root_p).parts) > 1 else "",
                    "sha256": "",
                    "pdf_pages": "", "pdf_encrypted": "", "pdf_scanned": "",
                    "xlsx_sheets": "", "xlsx_max_rows": "", "xlsx_has_formulas": "",
                }
                if do_hash and st.st_size > 0:
                    try:
                        rec["sha256"] = sha256(p)
                    except OSError:
                        pass
                if ext == "pdf" and st.st_size > 0:
                    rec.update(pdf_hazards(p))
                elif ext in ("xlsx", "xlsm") and st.st_size > 0:
                    rec.update(xlsx_hazards(p))
                yield rec


def summarize(rows: list[dict]) -> dict:
    by_family = defaultdict(lambda: {"files": 0, "bytes": 0})
    by_ext = defaultdict(lambda: {"files": 0, "bytes": 0})
    by_year = defaultdict(lambda: {"files": 0, "bytes": 0})
    by_top = defaultdict(lambda: {"files": 0, "bytes": 0})
    for r in rows:
        for key, bucket in ((r["family"], by_family), (r["ext"] or "(none)", by_ext),
                            (r["year"], by_year), (f'{r["root"]}/{r["top_folder"]}', by_top)):
            bucket[key]["files"] += 1
            bucket[key]["bytes"] += r["size_bytes"]
    hashes = Counter(r["sha256"] for r in rows if r["sha256"])
    dup_files = sum(c - 1 for c in hashes.values() if c > 1)
    dup_bytes = sum(next(r["size_bytes"] for r in rows if r["sha256"] == h) * (c - 1)
                    for h, c in hashes.items() if c > 1)
    hazards = {
        "scanned_pdfs": sum(1 for r in rows if r["pdf_scanned"] == "yes"),
        "encrypted_pdfs": sum(1 for r in rows if r["pdf_encrypted"] == "yes"),
        "pdf_parse_errors": sum(1 for r in rows if str(r["pdf_encrypted"]).startswith("error")),
        "legacy_office": sum(1 for r in rows if r["family"].endswith("-legacy")),
        "multi_sheet_workbooks": sum(1 for r in rows if isinstance(r["xlsx_sheets"], int) and r["xlsx_sheets"] > 1),
        "workbooks_with_formulas": sum(1 for r in rows if r["xlsx_has_formulas"] == "yes"),
        "zero_byte_files": sum(1 for r in rows if r["size_bytes"] == 0),
        "compressed_archives": sum(1 for r in rows if r["family"] == "archive"),
        "outlook_pst": sum(1 for r in rows if r["family"] == "email-pst"),
        "db_dumps": sum(1 for r in rows if r["family"] == "db-dump"),
        "files_over_50mb": sum(1 for r in rows if r["size_bytes"] > 50 * 1024 * 1024),
    }
    years = sorted(int(y) for y in by_year)
    return {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "total_files": len(rows),
        "total_bytes": sum(r["size_bytes"] for r in rows),
        "year_range": [years[0], years[-1]] if years else None,
        "duplicates": {"duplicate_files": dup_files, "duplicate_bytes": dup_bytes,
                       "note": "hash-based; requires --hash (default on)"},
        "hazards": hazards,
        "by_family": dict(sorted(by_family.items(), key=lambda kv: -kv[1]["bytes"])),
        "by_ext": dict(sorted(by_ext.items(), key=lambda kv: -kv[1]["files"])),
        "by_year": dict(sorted(by_year.items())),
        "by_top_folder": dict(sorted(by_top.items(), key=lambda kv: -kv[1]["bytes"])[:50]),
    }


def pick_samples(rows: list[dict], n: int) -> list[dict]:
    """Worst-case files per family: scanned/encrypted/legacy first, then largest, then oldest."""
    def score(r: dict) -> tuple:
        hazard = 0
        if r["pdf_scanned"] == "yes": hazard += 3
        if r["pdf_encrypted"] == "yes" or str(r["pdf_encrypted"]).startswith("error"): hazard += 3
        if r["family"].endswith("-legacy"): hazard += 2
        if isinstance(r["xlsx_sheets"], int) and r["xlsx_sheets"] > 3: hazard += 2
        if r["xlsx_has_formulas"] == "yes": hazard += 1
        if isinstance(r["pdf_pages"], int) and r["pdf_pages"] > 100: hazard += 1
        return (-hazard, -r["size_bytes"], r["year"])
    out = []
    by_family = defaultdict(list)
    for r in rows:
        if r["family"] in ("binary", "archive", "image", "other", "no-extension", "db-dump"):
            continue
        by_family[r["family"]].append(r)
    for fam, items in sorted(by_family.items()):
        items.sort(key=score)
        # half worst-case, half random spread across years for realism
        worst = items[: max(1, n // 2)]
        rest = [r for r in items if r not in worst]
        step = max(1, len(rest) // max(1, n - len(worst)))
        spread = rest[::step][: n - len(worst)]
        for r in worst + spread:
            reason = []
            if r["pdf_scanned"] == "yes": reason.append("scanned")
            if r["pdf_encrypted"] == "yes": reason.append("encrypted")
            if str(r["pdf_encrypted"]).startswith("error"): reason.append("parse-error")
            if r["family"].endswith("-legacy"): reason.append("legacy-format")
            if isinstance(r["xlsx_sheets"], int) and r["xlsx_sheets"] > 3: reason.append(f'{r["xlsx_sheets"]}-sheets')
            if r["xlsx_has_formulas"] == "yes": reason.append("formulas")
            if r["size_bytes"] > 20 * 1024 * 1024: reason.append("large")
            out.append({"family": fam, "path": r["path"], "size_bytes": r["size_bytes"], "year": r["year"],
                        "reason": ",".join(reason) or "representative",
                        "parser_result": "", "notes": ""})
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("roots", nargs="+", help="folders to inventory")
    ap.add_argument("--out", default="census_out")
    ap.add_argument("--no-hash", action="store_true", help="skip SHA-256 (no duplicate detection)")
    ap.add_argument("--samples", type=int, default=20, help="samples per family for parser testing")
    ap.add_argument("--follow-links", action="store_true")
    args = ap.parse_args()

    out = Path(args.out); out.mkdir(parents=True, exist_ok=True)
    t0 = time.time()
    rows: list[dict] = []
    for i, rec in enumerate(walk(args.roots, not args.no_hash, args.follow_links), 1):
        rows.append(rec)
        if i % 2000 == 0:
            print(f"  {i} files, {time.time()-t0:.0f}s", file=sys.stderr)
    if not rows:
        print("no files found", file=sys.stderr); sys.exit(1)

    with (out / "files.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)
    summary = summarize(rows)
    (out / "files_summary.json").write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    samples = pick_samples(rows, args.samples)
    with (out / "samples.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(samples[0].keys()) if samples else ["path"]); w.writeheader(); w.writerows(samples)

    gb = summary["total_bytes"] / 1e9
    print(f"\n{summary['total_files']:,} files, {gb:.2f} GB, years {summary['year_range']}, "
          f"{summary['duplicates']['duplicate_files']:,} duplicate files ({summary['duplicates']['duplicate_bytes']/1e9:.2f} GB)")
    print("hazards:", json.dumps(summary["hazards"]))
    print("by family:")
    for fam, v in summary["by_family"].items():
        print(f"  {fam:14s} {v['files']:>8,} files {v['bytes']/1e9:8.2f} GB")
    print(f"\nwrote {out/'files.csv'}, {out/'files_summary.json'}, {out/'samples.csv'} ({len(samples)} samples) in {time.time()-t0:.0f}s")


if __name__ == "__main__":
    main()
