#!/usr/bin/env python3
"""
run_eval.py — score RAGFlow retrieval against the Phase 0 evaluation set.

For every question in evaluation_set.xlsx (sheet 'Questions') it calls RAGFlow's
/retrieval endpoint across the BridgeIT knowledge bases, then checks:
  Hit@5  — is the expected source file among the top-5 chunks' documents?
           (matched by file name taken from the 'Expected source' cell, case-insensitive)
  Rank   — position of the first hit (1–5)
  Exact  — does any top-5 chunk contain the 'Expected answer snippet' verbatim
           (whitespace-normalised, case-insensitive)?
Results are written to eval_results_<label>.csv and, unless --no-xlsx, into the
next free run block of sheet 'Runs' in the workbook, where Recall@5 / MRR / exact-match
formulas already exist.

Usage:
  python run_eval.py --config config.local.yaml --xlsx ../phase0/evaluation_set.xlsx --label "baseline"
  python run_eval.py --config config.local.yaml --xlsx evaluation_set.xlsx --label "hybrid+rerank" --rerank
  python run_eval.py ... --vector-weight 0.7        # semantic-heavy comparison
  python run_eval.py ... --kb bt-projb_2016         # restrict to one KB (repeatable)
  python run_eval.py ... --filter-by-project        # add metadata_condition project = 'Project / domain' cell
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from pathlib import Path

import openpyxl
import yaml

from ragflow_client import RagflowClient, RagflowError

FILE_RE = re.compile(r"([\w\-. ()&]+\.(?:pdf|docx?|xlsx?|xlsm|csv|pptx?|txt|md|eml|msg|html?|sql|java|ts|xml|json))", re.I)


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


def expected_file(src: str) -> str:
    m = FILE_RE.search(src or "")
    return m.group(1).strip().lower() if m else ""


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--config", default="config.yaml")
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--label", required=True, help="run label written into the Runs sheet header")
    ap.add_argument("--kb", action="append", default=[], help="KB name(s) to search; default = all with the config prefix")
    ap.add_argument("--rerank", action="store_true", help="use rerank_model from config")
    ap.add_argument("--vector-weight", type=float)
    ap.add_argument("--page-size", type=int)
    ap.add_argument("--filter-by-project", action="store_true")
    ap.add_argument("--no-xlsx", action="store_true")
    args = ap.parse_args()

    cfg = yaml.safe_load(Path(args.config).read_text(encoding="utf-8"))
    rf, ev = cfg["ragflow"], cfg["eval"]
    client = RagflowClient(rf["base_url"], rf["api_key"])
    prefix = cfg["knowledge_bases"].get("name_prefix", "")
    datasets = [d for d in client.list_datasets() if (d["name"] in args.kb if args.kb else d["name"].startswith(prefix))]
    if not datasets:
        print("no knowledge bases found", file=sys.stderr); sys.exit(1)
    ds_ids = [d["id"] for d in datasets]
    print(f"searching {len(ds_ids)} KBs: {', '.join(d['name'] for d in datasets)}")

    wb = openpyxl.load_workbook(args.xlsx)
    qs = wb["Questions"]
    questions = []
    for row in qs.iter_rows(min_row=2, values_only=True):
        qid, text = row[0], row[1]
        if not text:
            continue
        questions.append({"id": qid, "question": str(text), "qtype": row[2] or "", "project": row[4] or "",
                          "source": str(row[6] or ""), "snippet": str(row[7] or "")})
    if not questions:
        print("no questions in the workbook", file=sys.stderr); sys.exit(1)

    page_size = args.page_size or ev.get("page_size", 5)
    vw = args.vector_weight if args.vector_weight is not None else ev.get("vector_similarity_weight", 0.3)
    rerank_id = rf.get("rerank_model") if (args.rerank or ev.get("use_rerank")) and rf.get("rerank_model") else None

    results, t0 = [], time.time()
    for q in questions:
        cond = None
        if args.filter_by_project and q["project"]:
            cond = {"logic": "and", "conditions": [{"name": "project", "comparison_operator": "contains", "value": q["project"]}]}
        try:
            data = client.retrieve(q["question"], ds_ids, page_size=page_size,
                                   similarity_threshold=ev.get("similarity_threshold", 0.1),
                                   vector_similarity_weight=vw, top_k=ev.get("top_k", 1024),
                                   rerank_id=rerank_id, keyword=ev.get("keyword", True), metadata_condition=cond)
            chunks = data.get("chunks", [])
        except RagflowError as e:
            print(f"!! {q['id']}: {e}", file=sys.stderr); chunks = []
        exp_file, exp_snip = expected_file(q["source"]), norm(q["snippet"])
        rank, exact, top_docs = 0, 0, []
        for i, c in enumerate(chunks[:page_size], 1):
            docname = (c.get("document_keyword") or c.get("docnm_kwd") or "").lower()
            top_docs.append(docname)
            # display names carry the relative path with '__' separators; compare on the file name only
            if exp_file and not rank and (exp_file in docname or docname.split("__")[-1] == exp_file):
                rank = i
            if exp_snip and exp_snip in norm(c.get("content", "")):
                exact = 1
        hit = 1 if rank else 0
        results.append({"id": q["id"], "question": q["question"], "qtype": q["qtype"], "expected_file": exp_file,
                        "hit@5": hit, "rank": rank or "", "exact": exact,
                        "top_docs": " | ".join(top_docs), "top_score": chunks[0].get("similarity", "") if chunks else "",
                        "top_content": (chunks[0].get("content", "")[:300] if chunks else "")})
        print(f"{q['id']}  hit={hit} rank={rank or '-'} exact={exact}  {q['question'][:70]}")

    n = len(results)
    recall = sum(r["hit@5"] for r in results) / n
    mrr = sum(1 / r["rank"] for r in results if r["rank"]) / n
    ex_rate = sum(r["exact"] for r in results) / n
    print(f"\n[{args.label}] n={n}  Recall@{page_size}={recall:.1%}  MRR={mrr:.3f}  Exact={ex_rate:.1%}  "
          f"(vector_weight={vw}, rerank={'on' if rerank_id else 'off'}) {time.time()-t0:.0f}s")
    by_type = {}
    for r in results:
        by_type.setdefault(r["qtype"] or "?", []).append(r)
    for t, rs in sorted(by_type.items()):
        print(f"   {t:14s} n={len(rs):3d} recall={sum(x['hit@5'] for x in rs)/len(rs):.0%} exact={sum(x['exact'] for x in rs)/len(rs):.0%}")

    out_csv = Path(f"eval_results_{re.sub(r'[^A-Za-z0-9]+', '_', args.label)}.csv")
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(results[0].keys())); w.writeheader(); w.writerows(results)
    print(f"wrote {out_csv}")

    if args.no_xlsx:
        return
    runs = wb["Runs"]
    # reuse the block whose header (row 2) equals the label; else the first block with no scores yet
    col = None
    for c in range(3, runs.max_column + 2, 3):
        v = runs.cell(row=2, column=c).value
        has_data = any(runs.cell(row=rr, column=c).value is not None for rr in range(4, 104))
        if v == args.label or not has_data:
            col = c; break
    if col is None:
        col = runs.max_column + 1
    runs.cell(row=2, column=col, value=args.label)
    for off, h in enumerate(["Hit@5", "Rank", "Exact"]):
        runs.cell(row=3, column=col + off, value=h)
    by_id = {r["id"]: r for r in results}
    for row in range(4, 104):
        qid = qs.cell(row=row - 2, column=1).value
        r = by_id.get(qid)
        if r:
            runs.cell(row=row, column=col, value=r["hit@5"])
            runs.cell(row=row, column=col + 1, value=r["rank"] or None)
            runs.cell(row=row, column=col + 2, value=r["exact"])
    # summary formulas for a new block
    from openpyxl.utils import get_column_letter as L
    hit, rank, ex = L(col), L(col + 1), L(col + 2)
    runs[f"{hit}106"] = f'=IF(COUNT({hit}4:{hit}103)=0,0,SUM({hit}4:{hit}103)/COUNT({hit}4:{hit}103))'
    runs[f"{hit}107"] = f'=IF(COUNT({hit}4:{hit}103)=0,0,SUMPRODUCT(({rank}4:{rank}103>0)/IF({rank}4:{rank}103>0,{rank}4:{rank}103,1))/COUNT({hit}4:{hit}103))'
    runs[f"{hit}108"] = f'=IF(COUNT({ex}4:{ex}103)=0,0,SUM({ex}4:{ex}103)/COUNT({ex}4:{ex}103))'
    runs[f"{hit}109"] = f'=COUNT({hit}4:{hit}103)'
    wb.save(args.xlsx)
    print(f"scores written to '{args.xlsx}' sheet Runs, block '{args.label}'")


if __name__ == "__main__":
    main()
